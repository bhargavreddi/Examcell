

__author__ = 'Bhargav'
import django
import sys,os
sys.path.append("/home/bhargav/Examcell")
sys.path.append("/home/mvgrexamcell/Examcell")
os.environ["DJANGO_SETTINGS_MODULE"] = "Examcell.settings"
django.setup()
from openpyxl import load_workbook
from Examination.models import Classroom, ClassroomCapacity, Department
workbook = load_workbook('classrooms.xlsx')
row = 2 
column = 1
sheet = workbook['Sheet1']
Classroom.objects.all().delete()
while True:
	cell = sheet.cell(row = row,column = column)
	if cell.value == None:
		break
	classroom = cell.value
	capacity = sheet.cell(row = row,column = column + 1).value
	department = sheet.cell(row = row,column = column + 2).value
	row = row + 1
	dept_names = {'cse': 'Computer Science and Engineering', 'eee': 'Electrical & Electronics Engineering','ece': 'Electronics and Communication Engineering', 'it': 'Information Technology','mech': 'Mechanical Engineering', 'chem': 'Chemical Engineering', 'civil': 'Civil Engineering'}
	obj = Classroom()
	obj.classroom_number = classroom
	obj.capacity = ClassroomCapacity.objects.get(size = capacity)
	dept = dept_names[department]
	obj.dept  = Department.objects.all().get(name = dept)
	obj.save()


