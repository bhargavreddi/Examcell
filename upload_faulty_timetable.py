__author__ = 'Bhargav'
import django
import sys,os
sys.path.append("/home/bhargav/Examcell")
sys.path.append("/home/mvgrexamcell/Examcell")
os.environ["DJANGO_SETTINGS_MODULE"] = "Examcell.settings"
django.setup()
from openpyxl import load_workbook
from Examination.models import Faculty
from Examination.models import FacultyTimetable
from Examination.models import Day
#Opening the workbook
workbook = load_workbook('2016-17 SEM-II FACULTY INDIVIDUALS- FINAL.xlsx')
sheet1 = workbook['Sheet1']
days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
FacultyTimetable.objects.all().delete()
row = 1
column = 1
while True:
	while True:
		cell = sheet1.cell(row = row,column = column)
		if cell.value != None:
			faculty_name = cell.value
			day = [None,None,None,None,None,None,None,None,None]
			for i in range(1,7):
				obj = FacultyTimetable()
				obj.faculty = Faculty.objects.get(faculty_name = faculty_name)
				obj.day = Day.objects.get(day = days[i-1])
				day[1] = sheet1.cell(row = row + i, column = column + 1).value
				if day[1] is None or day[1] == ' ':
					day[1] = 'Free'
				obj.hour1 = day[1]
				day[2] = sheet1.cell(row = row + i, column = column + 2).value
				if day[2] is None or day[2] == ' ':
					if sheet1.cell(row = row + i,column = column + 1).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 2).coordinate in sheet1.merged_cells:
						day[2] = day[1]
					else:
						day[2] = 'Free'
				obj.hour2 = day[2]
				day[3] = sheet1.cell(row = row + i, column = column + 3).value
				if day[3] is None or day[3] == ' ':
					if sheet1.cell(row = row + i,column = column + 2).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 3).coordinate in sheet1.merged_cells:
						day[3] = day[2]
					else:
						day[3] = 'Free'
				obj.hour3 = day[3]
				day[4] = sheet1.cell(row = row + i, column = column + 4).value
				if day[4] is None or day[4] == ' ':
					if sheet1.cell(row = row + i,column = column + 3).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 4).coordinate in sheet1.merged_cells:
						day[4] = day[3]
					else:
						day[4] = 'Free'
				obj.hour4 = day[4]
				day[5] = sheet1.cell(row = row + i, column = column + 5).value
				if day[5] is None or day[5] == ' ':
					if sheet1.cell(row = row + i,column = column + 4).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 5).coordinate in sheet1.merged_cells:
						day[5] = day[4]
					else:
						day[5] = 'Free'
				obj.hour5 = day[5]
				day[6] = sheet1.cell(row = row + i, column = column + 6).value
				if day[6] is None or day[6] == ' ':
					if sheet1.cell(row = row + i,column = column + 5).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 6).coordinate in sheet1.merged_cells:
						day[6] = day[5]
					else:
						day[6] = 'Free'
				obj.hour6 = day[6]
				day[7] = sheet1.cell(row = row + i, column = column + 7).value
				if day[7] is None or day[7] == ' ':
					if sheet1.cell(row = row + i,column = column + 6).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 7).coordinate in sheet1.merged_cells:
						day[7] = day[6]
					else:
						day[7] = 'Free'
				obj.hour7 = day[7]
				day[8] = sheet1.cell(row = row + i, column = column + 8).value
				if day[8] is None or day[8] == ' ':
					if sheet1.cell(row = row + i,column = column + 7).coordinate in sheet1.merged_cells and sheet1.cell(row = row + i,column = column + 8).coordinate in sheet1.merged_cells:
						day[8] = day[7]
					else:
						day[8] = 'Free'
				obj.hour8 = day[8]
				obj.save()
		else:
			row = row + 9
			column = 1
			break
		column = column + 10
	cell = sheet1.cell(row = row,column = column)
	if cell.value == None:
		break