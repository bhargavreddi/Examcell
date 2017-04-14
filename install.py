__author__ = 'Bhargav'
import django
import sys,os
from openpyxl import load_workbook
#from django.core.management import setup_environ
#from MissionRnD import settings
sys.path.append("/home/mvgrexamcell/Examcell")
sys.path.append("/home/bhargav/Examcell")
os.environ["DJANGO_SETTINGS_MODULE"] = "Examcell.settings"
django.setup()
from Examination.models import Department,Regulation,ClassroomCapacity, Year, Semester, Months, ExaminationType, Student, \
    Day, Examination, Subject
from Examination.models import Faculty
cse = Department()
cse.name = 'Computer Science and Engineering'
cse.save()
mech = Department()
mech.name = 'Mechanical Engineering'
mech.save()
chem = Department()
chem.name = 'Chemical Engineering'
chem.save()
civil = Department()
civil.name = 'Civil Engineering'
civil.save()
ece = Department()
ece.name = 'Electronics and Communication Engineering'
ece.save()
eee = Department()
eee.name = 'Electrical & Electronics Engineering'
eee.save()
it = Department()
it.name = 'Information Technology'
it.save()
unknown_dept = Department()
unknown_dept.name = 'Unknown'
unknown_dept.save()
r13 = Regulation()
r13.regulation = 'R13'
r13.save()
r16 = Regulation()
r16.regulation = 'R16'
r16.save()
cc24 = ClassroomCapacity()
cc24.size = 24
cc24.save()
cc36 = ClassroomCapacity()
cc36.size = 36
cc36.save()
cc56 = ClassroomCapacity()
cc56.size = 56
cc56.save()
y1 = Year()
y1.year = 1
y1.save()
y2 = Year()
y2.year = 2
y2.save()
y3 = Year()
y3.year = 3
y3.save()
y4 = Year()
y4.year = 4
y4.save()
s1 = Semester()
s1.semester = 1
s1.save()
s2 = Semester()
s2.semester = 2
s2.save()
months = ['January' , 'Februrary','March','April','May','June','July','August','September','October','November','December']
for month in months:
    obj = Months()
    obj.month = month
    obj.save()
examinationType = ['Regular', 'Supply','Mid Term']
for type in examinationType:
    obj = ExaminationType()
    obj.type = type
    obj.save()
#Student List Upload
workbook = load_workbook(filename = 'CSE.xlsx')
sheet = workbook['Sheet1']
row = 2
while True:
    if sheet.cell(row=row, column=1).value == None:
        break
    regno = sheet.cell(row=row, column=2).value
    student_name = sheet.cell(row=row, column=3).value
    obj = Student()
    obj.student_id = regno
    obj.student_name = student_name
    obj.student_year = Year.objects.all().get(year='4')
    obj.student_semester = Semester.objects.all().get(semester='2')
    obj.regulation = Regulation.objects.all().get(regulation='R13')
    obj.dept = Department.objects.all().get(name='Computer Science and Engineering')
    obj.save()
    row = row + 1
#Faculty List Upload
workbook = load_workbook(filename = 'faculty.xlsx')
sheet = workbook['T']
row = 3
while True:
    if sheet.cell(row=row, column=1).value == None:
        break
    name = sheet.cell(row=row, column=2).value
    dept = sheet.cell(row=row, column=4).value
    id = sheet.cell(row=row,column=5).value
    if dept != 'CSE':
        row = row + 1
        continue
    obj = Faculty()
    obj.faculty_id = id
    obj.faculty_name = name
    obj.dept = Department.objects.all().get(name='Computer Science and Engineering')
    obj.email = 'bhargav2014duggu@gmail.com'
    obj.save()
    row = row + 1


days = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday']
for day in days:
    obj = Day()
    obj.day = day
    obj.save()


obj = Examination()
obj.examination_name = "4-2 II Mid Examination 2017"
obj.type = ExaminationType.objects.get(type = 'Mid Term')
obj.month = Months.objects.get(month='April')
obj.year = 2017
obj.save()

subjects = [('RT42051','Distributed Systems','R13',False),('RT42052','Management Science','R13',False),('RT42043E','Cloud Computing','R13',False),('RT42053A','Human Computer Interaction','R13',False)]
for subject in subjects:
    obj = Subject()
    obj.subject_code = subject[0]
    obj.subject_name = subject[1]
    obj.regulation = Regulation.objects.get(regulation=subject[2])
    obj.drawingType = False
    obj.save()

obj = Student()
obj.student_id = '1'
obj.student_name = 'NULL'
obj.student_semester = Semester.objects.get(semester=1)
obj.student_year = Year.objects.get(year=1)
obj.regulation = Regulation.objects.get(regulation='R13')
obj.dept = Department.objects.get(name='Unknown')
obj.save()
